# -*- coding: utf-8 -*-
"""
Created on Wed May 17 13:11:01 2023

@author: LIPS
"""

# Importing libraries
# --------------------------------------------------------------------
import os
import pickle
import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

# Params
# --------------------------------------------------------------------
scale = "plastic_mix_500_1088"

# Directories
# --------------------------------------------------------------------
# getting the current directory
curr_dir = os.getcwd()

# Initial setup 
# --------------------------------------------------------------------
load_history = True

# getting parent directories 
parent_dir_1 = os.path.dirname(curr_dir)

# the folder path to save the model
saved_model = os.path.join(curr_dir, scale + "\models")

# the folder to save the history
saved_history = os.path.join(curr_dir, scale + "\history")

title = "(p,res=1088,manet,regnetx_160,adam,dice,d=500,c=50,t=0,lr=1e-5,e=50)"

# history files dir
train_loss_history = os.path.join(saved_history, "t_losses_" + title +".pkl")
valid_loss_history = os.path.join(saved_history, "v_losses_" + title +".pkl")
iou_score_history = os.path.join(saved_history, "iou_score_" + title +".pkl")
pixel_accuracy_history = os.path.join(saved_history, "accuracy_" + title +".pkl")
f1_score_history = os.path.join(saved_history, "f1score_" + title +".pkl")
precision_history = os.path.join(saved_history, "precision_" + title +".pkl")
recall_history = os.path.join(saved_history, "recall_" + title +".pkl")

# loading history
if load_history:
    pixel_accuracy = pickle.load(open(pixel_accuracy_history, "rb"))
    f1_score_history = pickle.load(open(f1_score_history, "rb"))
    iou_score = pickle.load(open(iou_score_history, "rb"))
    precision_history = pickle.load(open(precision_history, "rb"))
    recall_history = pickle.load(open(recall_history, "rb"))
    training_losses = pickle.load(open(train_loss_history, "rb"))
    validation_losses = pickle.load(open(valid_loss_history, "rb"))
    
    lst = []
    lst.append(pixel_accuracy[49])
    lst.append(f1_score_history[49])
    lst.append(iou_score[49])
    lst.append(precision_history[49])
    lst.append(recall_history[49])
    
    print(lst)

wb = load_workbook("C:/Users/GPU/particle_size_project/psd_metrics_study-filtered.xlsx")

ws = wb.active

i = 0

for col in range(13,18):
    char = get_column_letter(col)
    ws[char+str(109)].value = lst[i]
    i = i + 1

# saving 
wb.save("C:/Users/GPU/particle_size_project/psd_metrics_study-filtered.xlsx")